import openpyxl

def generate_sql():
    try:
        wb = openpyxl.load_workbook('اسماء المراكز والمحطات.xlsx', data_only=True)
        sql_lines = []
        sql_lines.append("-- Regions and Centers Seeding Script")
        sql_lines.append("-- Generated from 'اسماء المراكز والمحطات.xlsx'")
        sql_lines.append("")

        region_code = 1
        for sn in wb.sheetnames:
            ws = wb[sn]
            region_name = sn.strip()
            
            # 1. Insert Region
            # Using MERGE to be idempotent
            sql_lines.append(f"-- Region: {region_name}")
            sql_lines.append(f"MERGE INTO regions r USING (SELECT '{region_name}' n, '{region_code}' c FROM DUAL) src ")
            sql_lines.append(f"ON (r.name = src.n) ")
            sql_lines.append(f"WHEN MATCHED THEN UPDATE SET code = src.c ")
            sql_lines.append(f"WHEN NOT MATCHED THEN INSERT (name, code) VALUES (src.n, src.c);")
            sql_lines.append("")

            # 2. Extract Centers
            found_header = False
            center_idx = 1 # Serial index for centers in this region
            for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
                # Search for header row
                if not found_header:
                    if any(str(cell).strip() == 'المركز / المحطة' for cell in row if cell):
                        found_header = True
                        # Find column index for Center Name
                        try:
                            # It's usually column 2 (index 1)
                            for i, val in enumerate(row):
                                if str(val).strip() == 'المركز / المحطة':
                                    name_col = i
                                    break
                        except:
                            name_col = 1
                        continue
                else:
                    # Data row
                    center_serial = row[0] if row[0] is not None else center_idx
                    center_name = row[name_col] if name_col < len(row) and row[name_col] is not None else None
                    
                    if center_name:
                        center_name = str(center_name).strip()
                        # Use serial number directly as code
                        center_code_val = str(center_serial).strip()
                        
                        sql_lines.append(f"MERGE INTO centers c USING (SELECT '{center_name}' n, '{center_code_val}' c, (SELECT id FROM regions WHERE name = '{region_name}') rid FROM DUAL) src ")
                        sql_lines.append(f"ON (c.name = src.n) ")
                        sql_lines.append(f"WHEN MATCHED THEN UPDATE SET code = src.c, region_id = src.rid ")
                        sql_lines.append(f"WHEN NOT MATCHED THEN INSERT (name, code, region_id) VALUES (src.n, src.c, src.rid);")
                        center_idx += 1
            
            sql_lines.append("")
            region_code += 1

        sql_lines.append("COMMIT;")
        
        with open('seed_locations.sql', 'w', encoding='cp1256', errors='replace') as f:
            f.write("\n".join(sql_lines))
        
        print(f"Successfully generated seed_locations.sql with {len(sql_lines)} lines.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    generate_sql()
