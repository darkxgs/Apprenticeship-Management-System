import openpyxl
import sys

def check_excel():
    try:
        wb = openpyxl.load_workbook('اسماء المراكز والمحطات.xlsx', data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f"\n--- Sheet: {sn} ---")
            found_header = False
            for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
                # Look for a row that has 'المركز / المحطة' or similar
                if any(str(cell).strip() == 'المركز / المحطة' or str(cell).strip() == 'م' for cell in row if cell):
                    print(f"Found potential header row: {row}")
                    found_header = True
                    continue
                if found_header:
                    if any(row):
                         print(f"Data row: {row}")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    check_excel()
